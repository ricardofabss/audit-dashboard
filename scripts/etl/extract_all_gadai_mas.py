"""
extract_all_gadai_mas.py
Extracts ALL booking & settlement events from the Gadai MAS Excel files in a single pass.
Much faster than extract_gadai_mas_weekly.py because it only opens each Excel file once.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


def parse_date(value: Any) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, dt.datetime):
                return parsed.date()
            if isinstance(parsed, dt.date):
                return parsed
        except Exception:
            return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        formats = [
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%Y/%m/%d",
            "%d %b %Y",
            "%d %B %Y",
        ]
        for fmt in formats:
            try:
                return dt.datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    return None


def parse_datetime(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time(0, 0, 0)).isoformat()
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, dt.datetime):
                return parsed.isoformat()
            if isinstance(parsed, dt.date):
                return dt.datetime.combine(parsed, dt.time(0, 0, 0)).isoformat()
        except Exception:
            return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%d-%m-%Y %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d-%m-%Y %H:%M",
            "%d/%m/%Y %H:%M",
            "%Y-%m-%dT%H:%M:%S",
        ]
        for fmt in formats:
            try:
                return dt.datetime.strptime(raw, fmt).isoformat()
            except ValueError:
                continue
    return None


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        normalized = re.sub(r"\s+", "", raw)
        if "," in normalized and "." in normalized:
            normalized = normalized.replace(",", "")
        elif "," in normalized:
            normalized = normalized.replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def parse_int(value: Any) -> int | None:
    numeric = parse_number(value)
    if numeric is None:
        return None
    return int(round(numeric))


def safe_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def collateral_type_from_sheet(sheet_name: str) -> str:
    lowered = sheet_name.lower()
    if "elektronik" in lowered:
        return "ELEKTRONIK"
    return "EMAS"


def row_getter(header_map: dict[str, int]):
    def _get(row: tuple[Any, ...], name: str) -> Any:
        clean_name = re.sub(r"[\s_]+", "", name).lower() if name else ""
        idx = header_map.get(clean_name)
        if idx is None:
            return None
        if idx >= len(row):
            return None
        return row[idx]
    return _get


def get_any(getter, row: tuple[Any, ...], names: list[str]) -> Any:
    for name in names:
        value = getter(row, name)
        if value is not None and (not isinstance(value, str) or value.strip() != ""):
            return value
    return None


def extract_all_booking_events(path: Path) -> list[dict[str, Any]]:
    """Extract ALL booking events from the Excel file in a single pass."""
    events: list[dict[str, Any]] = []
    print(f"  Opening {path.name} ({path.stat().st_size / 1024 / 1024:.1f} MB)...", flush=True)
    wb = load_workbook(path, read_only=True, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            continue
        header_map = {}
        for idx, col in enumerate(header):
            if col is not None:
                clean_name = re.sub(r"[\s_]+", "", str(col)).lower()
                header_map[clean_name] = idx
        get = row_getter(header_map)
        collateral_type = collateral_type_from_sheet(sheet_name)
        sheet_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            contract_no = safe_str(get(row, "noSbg"))
            if not contract_no:
                continue

            disbursement_date = parse_date(get(row, "tglCair"))
            if not disbursement_date:
                continue
            register_date = parse_date(get(row, "tglRegister"))
            event_date = disbursement_date

            parent_contract_no = safe_str(get(row, "noSbgLama"))
            renewal_count = parse_int(get(row, "perpanjanganKe")) or 0
            is_renewal = renewal_count > 0 or bool(parent_contract_no)
            event_type = "BOOKING_RENEWAL" if is_renewal else "BOOKING_NEW"
            source_event_key = (
                f"BOOKING|{sheet_name}|{contract_no}|{event_date.isoformat()}|{row_idx}"
            )

            cif = safe_str(get(row, "cif"))
            cust_name = safe_str(get_any(get, row, ["namaNasabah", "namaCustomer", "nasabah", "customerName", "nama"]))
            cust_id_val = f"{cif} | {cust_name}" if cif and cust_name else cif

            events.append(
                {
                    "businessUnit": "GADAI_MAS",
                    "collateralType": collateral_type,
                    "contractNo": contract_no,
                    "parentContractNo": parent_contract_no,
                    "rootContractNo": parent_contract_no or contract_no,
                    "customerId": cust_id_val,
                    "outletCode": safe_str(get(row, "kodeOutlet")),
                    "outletName": safe_str(get(row, "namaOutlet")),
                    "branchName": safe_str(get(row, "namaCabang")),
                    "regionName": safe_str(get(row, "namaWilayah")),
                    "areaName": safe_str(get(row, "namaArea")),
                    "eventType": event_type,
                    "eventDate": event_date.isoformat(),
                    "eventTs": parse_datetime(get(row, "tanggalJam")),
                    "registerDate": register_date.isoformat() if register_date else None,
                    "disbursementDate": disbursement_date.isoformat() if disbursement_date else None,
                    "dueDate": None,
                    "settlementDate": parse_date(
                        get_any(get, row, ["tanggalPelunasan", "tglPelunasan", "tglPembayaran", "Tgl Pembayaran"])
                    ).isoformat()
                    if parse_date(get_any(get, row, ["tanggalPelunasan", "tglPelunasan", "tglPembayaran", "Tgl Pembayaran"]))
                    else None,
                    "tenorDays": parse_int(get(row, "tenor")),
                    "overdueDays": parse_int(get(row, "ovd")),
                    "renewalCount": renewal_count,
                    "isRenewal": is_renewal,
                    "ltvRatio": parse_number(get(row, "ltv")),
                    "principalInitial": parse_number(get(row, "pokokAwal")),
                    "loanInitial": parse_number(get(row, "pinjamAwal")),
                    "principalOutstanding": parse_number(get(row, "saldoPokok")),
                    "interestOutstanding": parse_number(get(row, "saldoBunga")),
                    "settlementAmount": None,
                    "saleAmount": None,
                    "interestIncome": None,
                    "settlementStatus": safe_str(get(row, "statusPerpanjangan")),
                    "exitStatus": None,
                    "sourceSystem": "BOOKING",
                    "sourceSheet": sheet_name,
                    "sourceEventKey": source_event_key,
                }
            )
            sheet_count += 1

        print(f"    Sheet '{sheet_name}': {sheet_count} booking events", flush=True)
    
    wb.close()
    return events


def extract_all_settlement_events(path: Path) -> list[dict[str, Any]]:
    """Extract ALL settlement events from the Excel file in a single pass."""
    events: list[dict[str, Any]] = []
    print(f"  Opening {path.name} ({path.stat().st_size / 1024 / 1024:.1f} MB)...", flush=True)
    wb = load_workbook(path, read_only=True, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            continue
        header_map = {}
        for idx, col in enumerate(header):
            if col is not None:
                clean_name = re.sub(r"[\s_]+", "", str(col)).lower()
                header_map[clean_name] = idx
        get = row_getter(header_map)
        collateral_type = collateral_type_from_sheet(sheet_name)
        sheet_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            contract_no = safe_str(get(row, "noSbg"))
            if not contract_no:
                continue

            settlement_date = parse_date(get_any(get, row, ["tanggalPelunasan", "tglPelunasan", "tglPembayaran", "Tgl Pembayaran"]))
            if not settlement_date:
                continue

            source_event_key = (
                f"PELUNASAN|{sheet_name}|{contract_no}|{settlement_date.isoformat()}|{row_idx}"
            )
            disbursement_date = parse_date(get(row, "tglCair"))
            due_date = parse_date(get(row, "tglJatuhTempo"))

            cif = safe_str(get(row, "cif"))
            cust_name = safe_str(get_any(get, row, ["namaNasabah", "namaCustomer", "nasabah", "customerName", "nama"]))
            cust_id_val = f"{cif} | {cust_name}" if cif and cust_name else cif

            events.append(
                {
                    "businessUnit": "GADAI_MAS",
                    "collateralType": collateral_type,
                    "contractNo": contract_no,
                    "parentContractNo": None,
                    "rootContractNo": contract_no,
                    "customerId": cust_id_val,
                    "outletCode": safe_str(get(row, "kodeOutlet")),
                    "outletName": safe_str(get(row, "namaOutlet")),
                    "branchName": safe_str(get(row, "namaCabang")),
                    "regionName": safe_str(get(row, "namaWilayah")),
                    "areaName": safe_str(get(row, "namaArea")),
                    "eventType": "SETTLEMENT",
                    "eventDate": settlement_date.isoformat(),
                    "eventTs": None,
                    "registerDate": None,
                    "disbursementDate": disbursement_date.isoformat() if disbursement_date else None,
                    "dueDate": due_date.isoformat() if due_date else None,
                    "settlementDate": settlement_date.isoformat(),
                    "tenorDays": parse_int(get(row, "tenor")),
                    "overdueDays": parse_int(get(row, "ovd")),
                    "renewalCount": None,
                    "isRenewal": False,
                    "ltvRatio": None,
                    "principalInitial": parse_number(get(row, "pokokAwal")),
                    "loanInitial": parse_number(get(row, "pinjamAwal")),
                    "principalOutstanding": None,
                    "interestOutstanding": None,
                    "settlementAmount": parse_number(get(row, "nilaiLunas")),
                    "saleAmount": parse_number(
                        get_any(get, row, ["nilaiPenjualan", "HPS"])
                    ),
                    "interestIncome": parse_number(get(row, "pendapatanBunga")),
                    "settlementStatus": safe_str(get(row, "statusPelunasan")),
                    "exitStatus": safe_str(get(row, "statusExit")),
                    "sourceSystem": "PELUNASAN",
                    "sourceSheet": sheet_name,
                    "sourceEventKey": source_event_key,
                }
            )
            sheet_count += 1

        print(f"    Sheet '{sheet_name}': {sheet_count} settlement events", flush=True)
    
    wb.close()
    return events


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract ALL Gadai MAS events from Excel (single pass)")
    parser.add_argument("--data-dir", required=True)
    parser.add_argument("--output", required=True, help="Output JSON file path")
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    booking_path = data_dir / "Booking.xlsx"
    settlement_path = data_dir / "Pelunasan.xlsx"

    if not booking_path.exists():
        raise FileNotFoundError(f"Booking file not found: {booking_path}")
    if not settlement_path.exists():
        raise FileNotFoundError(f"Pelunasan file not found: {settlement_path}")

    print("=== Extracting ALL Gadai MAS events (single pass) ===", flush=True)
    
    events = []
    events.extend(extract_all_booking_events(booking_path))
    print(f"  Total booking events: {len(events)}", flush=True)
    
    settlement_events = extract_all_settlement_events(settlement_path)
    events.extend(settlement_events)
    print(f"  Total settlement events: {len(settlement_events)}", flush=True)
    print(f"  Grand total events: {len(events)}", flush=True)

    # Find date range
    dates = [e["eventDate"] for e in events if e.get("eventDate")]
    if dates:
        print(f"  Date range: {min(dates)} to {max(dates)}", flush=True)

    payload = json.dumps({"events": events}, ensure_ascii=False)
    Path(args.output).write_text(payload, encoding="utf-8")
    print(f"  Written to: {args.output}", flush=True)


if __name__ == "__main__":
    main()
