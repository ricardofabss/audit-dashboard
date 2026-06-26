"""Peek at Excel headers only - writes results to a text file."""
import sys
sys.path.insert(0, '.')
from openpyxl import load_workbook
from pathlib import Path

DATA_DIR = Path(r"C:\Users\USER\OneDrive\Desktop\Kaizen\Sample data\Gadai MAS")
OUTPUT = Path(r"C:\Users\USER\OneDrive\Desktop\Kaizen\Ruang kerja audit\scripts\excel_headers_output.txt")

lines = []
for filename in ["Booking.xlsx", "Pelunasan.xlsx"]:
    p = DATA_DIR / filename
    if not p.exists():
        lines.append(f"NOT FOUND: {p}")
        continue
    lines.append(f"\n{'='*60}")
    lines.append(f"FILE: {filename} ({p.stat().st_size / 1024 / 1024:.1f} MB)")
    lines.append(f"{'='*60}")
    wb = load_workbook(p, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            lines.append(f"  Sheet '{sn}': NO HEADER")
            continue
        non_null = [h for h in header if h is not None]
        lines.append(f"\n  Sheet: '{sn}' - {len(non_null)} columns")
        for idx, col in enumerate(header):
            if col is not None:
                lines.append(f"    [{idx}] {col}")
        # Sample row
        sample = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if sample:
            lines.append(f"  --- Sample row ---")
            for idx, col in enumerate(header):
                if col is not None and idx < len(sample) and sample[idx] is not None:
                    val = str(sample[idx])[:80]
                    lines.append(f"    {col} = {val}")
    wb.close()

lines.append("\nDone!")
OUTPUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Output written to {OUTPUT}")
