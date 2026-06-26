import sys
from pathlib import Path
from openpyxl import load_workbook
import datetime as dt

data_dir = Path("C:\\Users\\USER\\OneDrive\\Desktop\\Kaizen\\Sample data\\Gadai MAS")
booking_path = data_dir / "Booking.xlsx"
settlement_path = data_dir / "Pelunasan.xlsx"

print("Checking sample files in:", data_dir)

if not booking_path.exists():
    print(f"Error: Booking.xlsx not found at {booking_path}")
    sys.exit(1)
if not settlement_path.exists():
    print(f"Error: Pelunasan.xlsx not found at {settlement_path}")
    sys.exit(1)

def get_sheet_date_range(file_path, date_columns):
    print(f"\nAnalyzing {file_path.name}...")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    min_date = None
    max_date = None
    row_count = 0
    
    for sheet_name in wb.sheetnames:
        print(f"  Reading sheet: {sheet_name}...")
        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            continue
            
        header_map = {str(col).strip(): idx for idx, col in enumerate(header) if col is not None}
        col_indices = [header_map[col] for col in date_columns if col in header_map]
        
        if not col_indices:
            print(f"    Warning: No date columns found in sheet {sheet_name}. Available: {list(header_map.keys())[:5]}...")
            continue
            
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            for idx in col_indices:
                if idx < len(row) and row[idx] is not None:
                    val = row[idx]
                    current_date = None
                    if isinstance(val, dt.datetime):
                        current_date = val.date()
                    elif isinstance(val, dt.date):
                        current_date = val
                    elif isinstance(val, str):
                        # Simple parse
                        val_clean = val.strip().split(" ")[0]
                        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
                            try:
                                current_date = dt.datetime.strptime(val_clean, fmt).date()
                                break
                            except ValueError:
                                continue
                    if current_date:
                        if min_date is None or current_date < min_date:
                            min_date = current_date
                        if max_date is None or current_date > max_date:
                            max_date = current_date
                            
    print(f"  Processed {row_count} rows.")
    print(f"  Date Range: {min_date} to {max_date}")
    return min_date, max_date

b_min, b_max = get_sheet_date_range(booking_path, ["tglCair", "tglRegister"])
s_min, s_max = get_sheet_date_range(settlement_path, ["tanggalPelunasan", "tglPelunasan"])

overall_min = min(filter(None, [b_min, s_min]))
overall_max = max(filter(None, [b_max, s_max]))

print("\n=== Summary of Sample Data ===")
print(f"Overall Date Range: {overall_min} to {overall_max}")

# Generate weekly suggestions
print("\n=== Recommended ETL Commands to Import the Data ===")
print("You can run the ETL tool manually by specifying dates. We suggest importing in weekly batches.")

# Let's break the range into weeks (1st-7th, 8th-14th, 15th-21st, 22nd-end)
current = dt.date(overall_min.year, overall_min.month, 1)
end_limit = overall_max

while current <= end_limit:
    year = current.year
    month = current.month
    
    # Week 1: 1-7
    w1_start = dt.date(year, month, 1)
    w1_end = dt.date(year, month, 7)
    if w1_start <= end_limit and w1_end >= overall_min:
        print(f"node scripts/etl/weekly_gadai_mas.mjs --windowStart {w1_start} --windowEnd {w1_end} --weekIndex 1 --periodMonth {month} --periodYear {year}")
        
    # Week 2: 8-14
    w2_start = dt.date(year, month, 8)
    w2_end = dt.date(year, month, 14)
    if w2_start <= end_limit and w2_end >= overall_min:
        print(f"node scripts/etl/weekly_gadai_mas.mjs --windowStart {w2_start} --windowEnd {w2_end} --weekIndex 2 --periodMonth {month} --periodYear {year}")
        
    # Week 3: 15-21
    w3_start = dt.date(year, month, 15)
    w3_end = dt.date(year, month, 21)
    if w3_start <= end_limit and w3_end >= overall_min:
        print(f"node scripts/etl/weekly_gadai_mas.mjs --windowStart {w3_start} --windowEnd {w3_end} --weekIndex 3 --periodMonth {month} --periodYear {year}")
        
    # Week 4: 22-end
    w4_start = dt.date(year, month, 22)
    # Get last day of month
    if month == 12:
        w4_end = dt.date(year + 1, 1, 1) - dt.timedelta(days=1)
    else:
        w4_end = dt.date(year, month + 1, 1) - dt.timedelta(days=1)
    if w4_start <= end_limit and w4_end >= overall_min:
        print(f"node scripts/etl/weekly_gadai_mas.mjs --windowStart {w4_start} --windowEnd {w4_end} --weekIndex 4 --periodMonth {month} --periodYear {year}")
        
    # Go to next month
    if month == 12:
        current = dt.date(year + 1, 1, 1)
    else:
        current = dt.date(year, month + 1, 1)
